"""
Code to run all scripts
"""
# Load in dependencies
import os
import tkinter
from tkinter import filedialog
import math
import os
import openpyxl as op

from AllResults import AllResults
from saveCurrentResults import saveCurrentResults
from BaseShear import BaseShear
from BraceDesign import BraceDesign
from BlockChecks import BlockChecks
from CorridorCurb import CorridorCurb
from CorridorContact import CorridorContact
from DiaphragmChecks import DiaphragmChecks
from IDR import IDR
from ElasticColumnDesign import ElasticColumnDesign

#call function to create all results CSVs
thisLoc, runName, groundMotion= AllResults()
# thisLoc = os.path.dirname(filedialog.askopenfilename())
# runName = 'r046_hs'
# groundMotion = 'REC03DC'

#create new Worksheet with results summary
resultsFolder, excelBook, excelSave = saveCurrentResults(runName, groundMotion)
# resultsFolder = ''
# excelSave = ''
# excelBook = op.load_workbook(excelSave)

#call function to perform Base Shear calcs and place values in summary sheet
maxX, minX, maxY, minY, Weight, SlidingCap = BaseShear(resultsFolder, thisLoc)
baseShear = excelBook["Base Shear"]
baseShear['B6'] = minX
baseShear['C6'] = maxX
baseShear['B7'] = minY
baseShear['C7'] = maxY
baseShear['E6'] = Weight
baseShear['E7'] = Weight
baseShear['F6'] = SlidingCap
baseShear['F7'] = SlidingCap
excelBook.save(excelSave)

# #call function to perform brace design and calcs
# maxDCR, maxTot, num_el = BraceDesign(thisLoc, resultsFolder)
# # maxDCR = maxDCR.tolist()
# braces = excelBook["Braces"]
# braces['B1'] = groundMotion
# capacity = 4*60*0.44
# for i in range(0,num_el):
    # DCR = float(maxDCR[i])
    # braces.cell(column = 2, row = i+6, value = maxDCR[i][0])
    # braces.cell(column = 4, row = i+6, value = maxTot[i][0])
    # braces.cell(column=5, row = i+6, value =capacity)
# excelBook.save(excelSave)

# #call function to perform bridging element calcs and add to summary sheet
# num_el = 9
# BlockCapX, BlockCapY, BlockCapComp, maxForceX, minForceX, maxForceY, minForceY = BlockChecks(thisLoc, resultsFolder)
# blocks = excelBook["Bridging Elements"]
# blocks['B1'] = groundMotion
# for i in range(0,num_el):
    # blocks.cell(column=2, row = i+6, value = maxForceX[i])
    # blocks.cell(column = 3, row = i+6, value = BlockCapX[i])
    # blocks.cell(column = 4, row = i+6, value = minForceX[i])
    # blocks.cell(column = 5, row = i+6, value = BlockCapComp[i])
    # blocks.cell(column = 6, row = i+6, value = maxForceY[i])
    # blocks.cell(column = 7, row = i+6, value = minForceY[i])
    # blocks.cell(column = 8, row= i+6, value = -BlockCapY[i])
# excelBook.save(excelSave)

# #call function to perform bridging element calcs for corridor curb
# maxT, maxC, maxShearY, minShearY, CurbTCap, CurbCCap, CurbShearCap = CorridorCurb(thisLoc, resultsFolder)
# curb = excelBook["Bridging Elements"]
# curb['B15'] = maxT
# curb['C15'] = CurbTCap
# curb['D15'] = maxC
# curb['E15'] = CurbCCap
# curb['F15'] = maxShearY
# curb['G15'] = minShearY
# curb['H15'] = CurbShearCap
# excelBook.save(excelSave)

# #call function to perform calculation at Unit4/Corridor Interface and add to summary sheets
# shearcap, tensionCap, maxT, maxC, maxShear, minShear = CorridorContact(thisLoc, resultsFolder)
# contact = excelBook["Bridging Elements"]
# contact['B16'] = maxT
# contact['C16'] = tensionCap
# contact['D16'] = maxC
# contact['E16'] = 490
# contact['F16'] = maxShear
# contact['G16'] = minShear
# contact['H16'] = shearcap
# excelBook.save(excelSave)

# #call function to perform diaphragm calculations and
# Capacities, Unit5RoofXMax, Unit5RoofXMin, Unit5RoofYMax, Unit5RoofYMin, Unit5L2XMax, Unit5L2XMin, Unit5L2YMax, Unit5L2YMin, Unit4L2XMax, Unit4L2XMin, Unit4L2YMax, Unit4L2YMin, CorridorLowRoofXMax, CorridorLowRoofXMin, CorridorLowRoofYMax, CorridorLowRoofYMin,CorridorRoofXMax, CorridorRoofXMin, CorridorRoofYMax, CorridorRoofYMin, CorridorL2XMax, CorridorL2XMin,CorridorL2YMax, CorridorL2YMin = DiaphragmChecks(thisLoc, resultsFolder)
# DiaphArray = [Unit5RoofXMax, Unit5RoofXMin, Unit5RoofYMax, Unit5RoofYMin,
             # Unit5L2XMax, Unit5L2XMin, Unit5L2YMax, Unit5L2YMin, Unit4L2XMax, Unit4L2XMin,
             # Unit4L2YMax, Unit4L2YMin, CorridorRoofXMax, CorridorRoofXMin, CorridorRoofYMax, CorridorRoofYMin,
             # CorridorLowRoofXMax, CorridorLowRoofXMin, CorridorLowRoofYMax,
             # CorridorLowRoofYMin, CorridorL2XMax,
             # CorridorL2XMin,CorridorL2YMax, CorridorL2YMin]

# num_objs = len(DiaphArray)
# diaphragm = excelBook["Diaphragms"]
# diaphragm['B1'] = groundMotion
# startRow = [8, 18,30,39, 50, 58, 70, 78, 85, 90, 97, 105]
# for k in range(0,len(Capacities)):
    # c = k*2
    # d = c + 1
    # for i in range(0,7):
        # if i < len(Capacities[k]):
            # diaphragm.cell(column = 2, row = startRow[k]+i, value = DiaphArray[c][i])
            # diaphragm.cell(column = 3, row = startRow[k]+i, value = DiaphArray[d][i])
            # diaphragm.cell(column = 4, row = startRow[k]+i, value = Capacities[k][i])
        # else:
            # break
# excelBook.save(excelSave)

#call function to perform drift calculations and add to summary sheet

#call function to perform Elastic Column Calculations
maxShearY, minShearY, maxShearZ, minShearZ = ElasticColumnDesign(thisLoc, resultsFolder)
elColumn = excelBook["Elastic Column"]
maxVy = max(maxShearY)
maxVz = max(maxShearZ)
minVy = min(minShearY)
minVz = min(minShearZ)
startRow = 6
elColumn.cell(column = 1, row = startRow, value = maxVy)
elColumn.cell(column = 2, row = startRow, value = minVy)
elColumn.cell(column = 3, row = startRow, value = maxVz)
elColumn.cell(column = 4, row = startRow, value = minVz)
excelBook.save(excelSave)
